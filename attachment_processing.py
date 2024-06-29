import cv2
import numpy as np
import pandas as pd
from paddleocr import PaddleOCR
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

# 设置OCR模型
ocr_model = PaddleOCR(lang='en', use_angle_cls=True)

def process_attachments_to_excel(date):
    # 此處添加你處理附件到Excel的程式碼
    # 定义要提取的区域坐标
    regions1 = {
        "region_1": ((411, 854), (650, 1050)),  # Rank
        "region_2": ((536, 420), (705, 455))  # 日期
    }
    regions2 = {
        "region_1": ((312, 312), (475, 386)),
        "region_2": ((312, 620), (475, 720)),
        "region_3": ((312, 935), (475, 1030)),
        "region_4": ((312, 1260), (475, 1330)),
        "region_5": ((312, 1580), (475, 1645)),
        "region_6": ((1800, 300), (1950, 400)),
        "region_7": ((1800, 620), (1950, 700)),
        "region_8": ((1800, 950), (1950, 1015)),
        "region_9": ((1800, 1260), (1950, 1330)),
        "region_10": ((1800, 1580), (1950, 1645)),
        "region_11": ((784, 405), (1385, 450)),
        "region_12": ((782, 723), (1385, 764)),
        "region_13": ((782, 1040), (1385, 1080)),
        "region_14": ((782, 1354), (1385, 1394)),
        "region_15": ((782, 1670), (1385, 1710)),
        "region_16": ((2270, 410), (2872, 450)),
        "region_17": ((2270, 725), (2872, 765)),
        "region_18": ((2270, 1040), (2872, 1090)),
        "region_19": ((2270, 1355), (2872, 1395)),
        "region_20": ((2254, 1670), (2872, 1710))
    }

    folder_path = f"motiphysio{date.replace('/', '')}"
    file_list = os.listdir(folder_path)

    grouped_files = {}
    for file_name in file_list:
        if file_name.endswith(".png"):
            parts = file_name.rsplit("_", 1)
            base_name = parts[0]
            suffix = parts[1].split(".")[0]

            if base_name not in grouped_files:
                grouped_files[base_name] = {}

            if suffix == "1":
                grouped_files[base_name]["img_path1"] = os.path.join(folder_path, file_name)
            elif suffix == "2":
                grouped_files[base_name]["img_path2"] = os.path.join(folder_path, file_name)

    for base_name, paths in grouped_files.items():
        if 'img_path1' in paths and 'img_path2' in paths:
            img_path1 = paths['img_path1']
            img_path2 = paths['img_path2']

        original_img1 = cv2.imdecode(np.fromfile(img_path1, dtype=np.uint8), cv2.IMREAD_COLOR)
        original_img2 = cv2.imdecode(np.fromfile(img_path2, dtype=np.uint8), cv2.IMREAD_COLOR)

        region_texts1 = {}
        region_texts2 = {}

        for region_name, ((x_start, y_start), (x_end, y_end)) in regions1.items():
            cropped_img = original_img1[y_start:y_end, x_start:x_end]
            result = ocr_model.ocr(cropped_img)
            collected_text = [element[1][0] for line in result for element in line]
            full_text = " ".join(collected_text)
            region_texts1[region_name] = full_text

        for i in range(1, 3):
            region_key = f"region_{i}"
            if region_key not in region_texts1:
                region_texts1[region_key] = ""

        for region_name, ((x_start, y_start), (x_end, y_end)) in regions2.items():
            cropped_img = original_img2[y_start:y_end, x_start:x_end]
            result = ocr_model.ocr(cropped_img)
            collected_text = [element[1][0] for line in result for element in line]
            full_text = " ".join(collected_text)
            region_texts2[region_name] = full_text

        for i in range(1, 21):
            region_key = f"region_{i}"
            if region_key not in region_texts2:
                region_texts2[region_key] = ""

        name = base_name
        region_descriptions = {}

        description_templates = {
            "region_1": {"Forward": "Head forward tilted", "Backward": "Head backward tilted"},
            "region_2": {"Right": "Right post. shoulder up", "Left": "Left post shoulder up"},
            "region_3": {"Right": "Right pelvic elevated", "Left": "Left pelvic elevated"},
            "region_4": {"Outward": "L't knee Outward tilted", "Inward": "L't knee Inward tilted"},
            "region_5": {"Outward": "R't knee Outward tilted", "Inward": "R't knee Inward tilted"},
            "region_6": {"Right": "Right ant. shoulder lifted", "Left": "Left ant. shoulder lifted"},
            "region_7": {"Forward": "Chest forward tilted", "Backward": "Chest backward tilted"},
            "region_8": {"Forward": "Pelvis forward move", "Backward": "Pelvis backward move"},
            "region_9": {"Right": "Pelvis right to left axial rotate", "Left": "Pelvis left to right axial rotate"},
            "region_10": {"Forward": "Knee forward flexed", "Backward": "Knee backward extended"}
        }

        for region, text in region_texts2.items():
            first_word = text.split()[0]
            if first_word in description_templates.get(region, {}):
                region_descriptions[region] = description_templates[region][first_word]
            else:
                region_descriptions[region] = "Unknown"

        column_data = {
            "Column_A": ["姓名", base_name],
            "Column_B": ["頭部位置", region_descriptions["region_1"]],
            "Column_C": ["", region_texts2["region_11"]],
            "Column_D": ["後側肩膀不平衡", region_descriptions["region_2"]],
            "Column_E": ["", region_texts2["region_12"]],
            "Column_F": ["骨盆傾斜", region_descriptions["region_3"]],
            "Column_G": ["", region_texts2["region_13"]],
            "Column_H": ["左膝內翻", region_descriptions["region_4"]],
            "Column_I": ["", region_texts2["region_14"]],
            "Column_J": ["右膝內翻", region_descriptions["region_5"]],
            "Column_K": ["", region_texts2["region_15"]],
            "Column_L": ["前側肩膀不平衡", region_descriptions["region_6"]],
            "Column_M": ["", region_texts2["region_16"]],
            "Column_N": ["上胸椎位置", region_descriptions["region_7"]],
            "Column_O": ["", region_texts2["region_17"]],
            "Column_P": ["骨盆前後位移", region_descriptions["region_8"]],
            "Column_Q": ["", region_texts2["region_18"]],
            "Column_R": ["骨盆軸向位移", region_descriptions["region_9"]],
            "Column_S": ["", region_texts2["region_19"]],
            "Column_T": ["膝屈伸", region_descriptions["region_10"]],
            "Column_U": ["", region_texts2["region_20"]],
            "Column_V": ["評級", region_texts1["region_1"]],
            "Column_W": ["檢測日期", region_texts1["region_2"]]
        }

        required_length = 2
        for key in column_data.keys():
            while len(column_data[key]) < required_length:
                column_data[key].append("")

        df = pd.DataFrame(column_data)
        excel_folder_path = os.path.join(folder_path, "Excel")
        os.makedirs(excel_folder_path, exist_ok=True)
        excel_path = os.path.join(excel_folder_path, f"{base_name}_detected_text.xlsx")
        df.to_excel(excel_path, index=False, header=False)

        wb = load_workbook(excel_path)
        ws = wb.active

        ws.merge_cells("B1:C1")
        ws.merge_cells("D1:E1")
        ws.merge_cells("F1:G1")
        ws.merge_cells("H1:I1")
        ws.merge_cells("J1:K1")
        ws.merge_cells("L1:M1")
        ws.merge_cells("N1:O1")
        ws.merge_cells("P1:Q1")
        ws.merge_cells("R1:S1")
        ws.merge_cells("T1:U1")

        columns_to_adjustA = ["B", "D", "F", "H", "J", "L", "N", "P", "R", "T"]
        for col in columns_to_adjustA:
            ws.column_dimensions[col].width = 20
        columns_to_adjustB = ["C", "E", "G", "I", "K", "M", "O", "Q", "S", "U"]
        for col in columns_to_adjustB:
            ws.column_dimensions[col].width = 6

        wb.save(excel_path)

        wb = load_workbook(excel_path)
        ws = wb.active

        numeric_columns = ["C", "E", "G", "I", "K", "M", "O", "Q", "S", "U"]
        numeric_values = {}

        for col in numeric_columns:
            cell_value = ws[f"{col}2"].value
            try:
                numeric_value = float(cell_value.strip('%'))
            except ValueError:
                numeric_value = 0
            numeric_values[col] = numeric_value

        sorted_columns = sorted(numeric_values.items(), key=lambda x: x[1], reverse=True)[:3]

        result_texts = []
        for col, value in sorted_columns:
            corresponding_col_letter = get_column_letter(column_index_from_string(col) - 1)
            corresponding_text = ws[f"{corresponding_col_letter}2"].value
            result_texts.append(corresponding_text)

        final_result = "; ".join(result_texts)

        lookup_excel_path = "C:/Users/lin/Documents/program/toss/Landseed/motiphysio_解釋對照表.xlsx"
        lookup_wb = load_workbook(lookup_excel_path)
        lookup_ws = lookup_wb.active

        sheet_name = "對應試算表"
        if sheet_name not in wb.sheetnames:
            ws2 = wb.create_sheet(sheet_name)
        else:
            ws2 = wb[sheet_name]

        row_index = 2
        for col, _ in sorted_columns:
            percentage_value = f"{numeric_values[col]}%"
            ws2[f"E{row_index}"] = percentage_value
            row_index += 1

        row_index = 2
        for corresponding_text in result_texts:
            ws2[f"D{row_index}"] = corresponding_text

            chinese_explain_f = ""
            chinese_explain_g = ""
            chinese_explain_h = ""
            chinese_explain_i = ""
            chinese_explain_j = ""
            chinese_explain_k = ""

            for row in lookup_ws.iter_rows(min_row=2, min_col=2, max_col=8):
                if row[0].value == corresponding_text:
                    chinese_explain_f = row[1].value
                    chinese_explain_g = row[2].value
                    chinese_explain_h = row[3].value
                    chinese_explain_i = row[4].value
                    chinese_explain_j = row[5].value
                    chinese_explain_k = row[6].value
                    break

            ws2[f"F{row_index}"] = chinese_explain_f
            ws2[f"G{row_index}"] = chinese_explain_g
            ws2[f"H{row_index}"] = chinese_explain_h
            ws2[f"I{row_index}"] = chinese_explain_i
            ws2[f"J{row_index}"] = chinese_explain_j
            ws2[f"K{row_index}"] = chinese_explain_k

            row_index += 1

        ws2["A1"] = "姓名"
        ws2["A2"] = name
        ws2["B1"] = "評級"
        ws2["B2"] = region_texts1["region_1"]
        ws2["C1"] = "檢測日期"
        ws2["C2"] = region_texts1["region_2"]
        ws2["D1"] = "前三項問題"
        ws2["E1"] = "風險"
        ws2["F1"] = "問題"
        ws2["G1"] = "問題描述"
        ws2["H1"] = "建議伸展動作"
        ws2["I1"] = "伸展敘述"
        ws2["J1"] = "建議訓練動作"
        ws2["K1"] = "訓練敘述"

        wb.save(excel_path)
